from styles import font, border_all, border_rls, border_top, alignment
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet("Bill", 0)

a1 = ws["A1"]
a1.value = "Hello"
a1.border = border_rls
a1.font = font
a1.alignment = alignment

wb.save('group.xlsx')

