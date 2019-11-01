from openpyxl import Workbook
import openpyxl
import os
from functions import back_from, goto
from values import sheet_start_from, path_excel_billing_folder
from styles import border_all as ba, font, alignment as am


class CreateNewExcelFile:
    def __init__(self, data_dictionary, data_len=8):
        if data_len < 7 or data_len <= 0:
            data_len = 8
        self.name = "_"
        self.data_len = data_len
        self.pos1 = sheet_start_from
        self.pos2 = self.pos1 + 2
        self.pos3 = self.pos2 + 1
        self.pos4 = self.pos3 + self.data_len
        # Create new workbook
        self.wb = Workbook()

        self.ws = self.wb.active

        # For Name & Address three row
        for name_row in range(self.pos1, self.pos2):
            if name_row == sheet_start_from:
                value = 'name'
            else:
                value = 'address'
            block = self.ws[f"B{name_row}"]
            block.value = f'{value.upper()}: {data_dictionary[value].title()}'
            self.ws.merge_cells(f"B{name_row}:E{name_row}")

        # For INVOICE number & INVOICE Date Row
        for invoice_row in range(self.pos1, self.pos2):
            if invoice_row == sheet_start_from:
                value = 'invoice no'
            else:
                value = 'date'
            block = self.ws[f"F{invoice_row}"]
            block.value = f'{value.upper()}: {data_dictionary[value]}'
            self.ws.merge_cells(f"F{invoice_row}:G{invoice_row}")

        # One row skipped
        # Heading

        sl_box = self.ws[f"A{self.pos3}"]
        sl_box.value = 'SL NO'
        sl_box.alignment = am
        prd_box = self.ws[f"B{self.pos3}"]
        prd_box.value = 'PRODUCT'
        prd_box.alignment = am
        price_box = self.ws[f"E{self.pos3}"]
        price_box.value = 'PRICE'
        price_box.alignment = am
        quantity_box = self.ws[f"F{self.pos3}"]
        quantity_box.value = 'QTY'
        quantity_box.alignment = am
        total_box = self.ws[f"G{self.pos3}"]
        total_box.value = 'TOTAL'
        total_box.alignment = am

        count = 1
        total = 0
        for sl in data_dictionary['products']:
            prd_row = self.pos3 + count
            serial = count
            product = sl[0].title()
            price = 350.00
            quantity = sl[1]
            prd_total = quantity * price
            total += prd_total
            count += 1
            sl_box = self.ws[f"A{prd_row}"]
            sl_box.value = serial
            sl_box.alignment = am
            prd_box = self.ws[f"B{prd_row}"]
            prd_box.value = product
            price_box = self.ws[f"E{prd_row}"]
            price_box.value = price
            quantity_box = self.ws[f"F{prd_row}"]
            quantity_box.value = quantity
            total_box = self.ws[f"G{prd_row}"]
            total_box.value = prd_total

        # For Product List with Heading
        for prd_row in range(self.pos3, self.pos4):
            self.ws.merge_cells(f"B{prd_row}:D{prd_row}")
            sl_b = self.ws[f"A{prd_row}"]
            sl_b.border = ba
            product_b = self.ws[f"B{prd_row}"]
            product_b.border = ba
            price_b = self.ws[f"E{prd_row}"]
            price_b.border = ba
            quantity_b = self.ws[f"F{prd_row}"]
            quantity_b.border = ba
            total_b = self.ws[f"G{prd_row}"]
            total_b.border = ba

        # To Save
        self.name = f"{data_dictionary['invoice no']} {data_dictionary['name']}-({data_dictionary['date']})"
        self.name = self.name.title().replace(' ', '_').replace('/', '.')
        goto(path_excel_billing_folder)
        self.wb.save(f"{self.name}.xlsx")
        back_from(path_excel_billing_folder)
